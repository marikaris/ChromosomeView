from openpyxl import load_workbook

class C6EmxFile:
    """Class EmxFile
    contains all the operations that an emx file can do (and are needed in this application)
    with the openpyxl library
    """
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(self.filename)
    
    def set_workbook(self):
        """function set_workbook
        sets the workbook value
        returns workbook (the workbook object)
        """
        workbook = load_workbook(self.filename)
        return workbook

    def set_sheet(self, sheet_name):
        """function set_sheet
        sets the sheet
        returns sheet, the sheet with the patient info
        """
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        return sheet

    def save_file_changes(self):
        """function save_file_changes
        saves the emx file with the original file name
        """
        self.workbook.save('new_'+self.filename[0:-1])
    
    def get_workbook(self):
        """function get_workbook
        returns the workbook to work with
        """
        return self.workbook
    
    def get_sheet(self):
        """function get_attr_sheet
        returns the attributes sheet
        """
        return self.sheet

    def alter_sheet(self, sheet, cell, value):
        """function alter_sheet
        This function can alter a sheet. It needs the sheet object to be altered,
        the cell which should be altered and the value that should be in the cell.
        """
        sheet[cell] = value

    def alter_cell(self, sheet, col, old_val, new_val, old_vals, new_vals):
        """function alter_cell
        This function alters the value of a cell in a given sheet.
        """
        columns = sheet.rows[0]
        for colNr, column in enumerate(columns):
            if column.value == col:
                for row in sheet.columns[colNr]:
                    values = str(row.value).split(',')
                    print(values)
                    newValues = []
                    #change the values that are altered
                    if values != [''] and values!= ['None']:
                        print('not none :',values)
                        for value in values:
                            if value == old_val:
                                newValues.append(new_val)
                                print(newValues)
                                print(old_val, new_val)
                            else:
                                if value not in old_vals or value not in new_vals:
                                    newValues.append(value)
                    if newValues != [''] and newValues!= ['None']:
                        print(newValues)
                        row.value = ','.join(newValues)

    def alter_attr_cell(self, sheet, old_val, new_val):
        """Function alter_attr_cell
        This function alters a cell in the attributes sheet.
        """
        columns = sheet.rows[0]
        for column in columns:
            if column.value == old_val:
                column.value = new_val

    def alter_attr_cells(self, altered_attributes, sheet):
        """Function alter_attr_cells
        Alters all attribute cells that should be altered, according to a given dictionary that contains
        the old values as keys and new values as values"""
        for attribute in altered_attributes:
            new_val = altered_attributes[attribute]
            self.alter_attr_cell(sheet, attribute, new_val)

    def alter_cells(self, altered_values, sheet):
        """Function alter_cells
        Alters all cells that should be altered, according to a given dictionary that contains all altered
        columns and the values (as keys) in that columns that are altered (dictionary like {oldval:newval}).
        """
        for column in altered_values:
            possible_values = altered_values[column]
            for value in possible_values:
                self.alter_cell(sheet, column, str(value), possible_values[str(value)], possible_values.keys(), possible_values.values())
