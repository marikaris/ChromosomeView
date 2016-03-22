from openpyxl import load_workbook


class EmxFile:
    """Class EmxFile
    contains all the operations that an emx file can do (and are needed in this application)
    with the openpyxl library
    """
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(self.filename)
        self.genes_sheet = self.set_genes_sheet()
        self.attr_sheet = self.set_attr_sheet()
        self.gwas_sheet = self.set_gwas_sheet()

    def set_workbook(self):
        """function set_workbook
        sets the workbook value
        returns workbook (the workbook object)
        """
        workbook = load_workbook(self.filename)
        return workbook
    
    def set_genes_sheet(self):
        """function set_genes_sheet
        sets the genes sheet value
        returns the genes_sheet
        """
        genes_sheet = self.workbook.create_sheet(title="genes")
        genes_sheet["A1"] = "ensembl_id"
        genes_sheet["B1"] = "gene_name"
        genes_sheet["C1"] = "omim_morbid_accesion"
        genes_sheet["D1"] = "omim_morbid_description"
        genes_sheet["E1"] = "start"
        genes_sheet["F1"] = "stop"
        genes_sheet["G1"] = "gene_info"
        genes_sheet["H1"] = "gene_type"
        genes_sheet["I1"] = "gwas"
        genes_sheet["J1"] = "cgd_condition"
        genes_sheet["K1"]  = "cgd_inheritance"
        return genes_sheet

    def set_gwas_sheet(self):
        gwas_sheet = self.workbook.create_sheet(title="gwas")
        gwas_sheet["A1"] = "gwas"
        gwas_sheet["B1"] = "pubmed"
        gwas_sheet["C1"] = "first_author"
        gwas_sheet["D1"] = "date"
        gwas_sheet["E1"] = "journal"
        gwas_sheet["F1"] = "link"
        gwas_sheet["G1"] = "study"
        gwas_sheet["H1"] = "reported_gene"
        gwas_sheet["I1"] = "mapped_gene"
        gwas_sheet["J1"] = "upstream_id"
        gwas_sheet["K1"] = "downstream_id"
        gwas_sheet["L1"] = "mapped_trait"
        return gwas_sheet

    def set_attr_sheet(self):
        """function set_attr_sheet
        sets the attribute sheet
        returns attrs, the sheet with the attributes table
        """
        attrs = self.workbook.get_sheet_by_name("attributes")
        return attrs

    def save_file_changes(self):
        """function save_file_changes
        saves the emx file with the original file name
        """
        self.workbook.save(self.filename)
    
    def get_workbook(self):
        """function get_workbook
        returns the workbook to work with
        """
        return self.workbook
    
    def get_attr_sheet(self):
        """function get_attr_sheet
        returns the attributes sheet
        """
        return self.attr_sheet
    
    def get_genes_sheet(self):
        """function get_genes_sheet
        returns the genes sheet
        """
        return self.genes_sheet

    def alter_sheet(self, sheet, cell, value):
        """function alter_sheet
        This function can alter a sheet. It needs the sheet object to be altered,
        the cell which should be altered and the value that should be in the cell.
        """
        sheet[cell] = value

    def find_first_empty_cell(self, sheet):
        """function find_first_empty_cell
        This function finds the first empty cell in the given sheet (checks if the first column is empty)
        Returns the row number of the cell."""
        first_col = "A"
        iteration = 1
        empty_cell_not_found = True
        while empty_cell_not_found:
            if sheet[first_col+str(iteration)].value== None:
                empty_cell_not_found= False
            else:
                iteration += 1
        return iteration

    def addAttr(self, attr_name, table_name, data_type, id_attr, nillable, description, ref_entity, default):
        """addAttr adds an attribute to the attribute table in the emx file. It needs all the values of the
        columns in the attributes table."""
        attr_sheet=self.attr_sheet
        empty_row = self.find_first_empty_cell(attr_sheet)
        col_list = ["A", "B", "C", "D", "E", "F", "G", "H"]
        item_list = [attr_name, table_name, data_type, id_attr, nillable, description, ref_entity, default]
        for col, val in zip(col_list, item_list):
            self.alter_sheet(attr_sheet, col+str(empty_row), val)

    def update_attr_sheet(self):
        self.addAttr("gene_name", "genes", "string", "FALSE", "FALSE", "the short name of the gene", "", "")
        self.addAttr("start", "genes", "long", "FALSE", "FALSE", "start position of the gene", "", "")
        self.addAttr("stop", "genes", "long", "FALSE", "FALSE", "stop position of the gene", "", "")
        self.addAttr("omim_morbid_accesion", "genes", "text", "FALSE", "TRUE", "The omim morbid accesion of the gene", "", "")
        self.addAttr("omim_morbid_description", "genes", "text", "FALSE", "TRUE", "The omim morbid description of the gene", "", "")
        self.addAttr("ensembl_id", "genes", "string", "TRUE","FALSE", "The ensembl accesion of the gene", "", "")
        self.addAttr("gene_info", "genes", "string", "FALSE", "TRUE", "The description of the gene", "", "")
        self.addAttr("gene_type", "genes", "string" ,"FALSE", "TRUE", "The type of the gene", "", "")
        self.addAttr("gwas", "genes", "mref", "FALSE", "TRUE", "The gwas articles that reported the gene.", "gwas", "")
        self.addAttr("cgd_condition", "genes", "text","FALSE", "TRUE", "The condition that is mapped to the gene in CGD", "", "")
        self.addAttr("cgd_inheritance", "genes", "string", "FALSE", "TRUE", "The inheritance that is expected of the trait mapped in CGD.", "", "")
        self.addAttr("gwas", "gwas", "string", "TRUE", "FALSE", "The id that links the study to the gene.", "", "")
        self.addAttr("pubmed", "gwas", "string", "FALSE", "FALSE", "The pubmed id of the gwas article.", "", "")
        self.addAttr("first_author", "gwas", "string", "FALSE", "TRUE", "The first author of the article", "", "")
        self.addAttr("date", "gwas", "string", "FALSE", "TRUE", "The publication date of the article", "", "")
        self.addAttr("journal", "gwas", "text", "FALSE", "TRUE", "The journal of the article", "", "")
        self.addAttr("link", "gwas", "text", "FALSE", "TRUE", "The link of the article", "", "")
        self.addAttr("study", "gwas", "text", "FALSE", "TRUE", "The name of the study", "", "")
        self.addAttr("reported_gene", "gwas", "text", "FALSE", "TRUE", "The reported gene(s) in the article", "", "")
        self.addAttr("mapped_gene", "gwas", "text", "FALSE", "TRUE", "The mapped gene(s) in the article", "", "")
        self.addAttr("downstream_id", "gwas", "text", "FALSE", "TRUE", "The downstream id reported in the article", "", "")
        self.addAttr("upstream_id", "gwas", "text", "FALSE", "TRUE", "The upstream id reported in the article", "", "")
        self.addAttr("mapped_trait", "gwas", "text", "FALSE", "TRUE", "The mapped trait in the article", "", "")

    def get_gwas_sheet(self):
        return self.gwas_sheet