from openpyxl import load_workbook

class EmxFile:
    """Class EmxFile
    contains all the operations that an emx file can do (and are needed in this application)
    with the openpyxl library
    """
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(self.filename)
        self.genes_sheet = self.set_genes_sheet()
        self.attr_sheet = self.set_attr_sheet()
    
    def set_workbook(self):
        """function set_workbook
        sets the workbook value
        returns workbook (the workbook object)
        """
        workbook = load_workbook(self.filename)
        return workbook
    
    def set_genes_sheet(self):
        """function set_genes_sheet
        sets the genes sheet value
        returns the genes_sheet
        """
        genes_sheet = self.workbook.create_sheet(title="genes")
        genes_sheet["A1"]="ensembl_id"
        genes_sheet["B1"]="gene_name"
        genes_sheet["C1"]="omim_morbid_accesion"
        genes_sheet["D1"]="omim_morbid_description"
        genes_sheet["E1"]="start"
        genes_sheet["F1"]="stop"
        return genes_sheet
    
    def set_attr_sheet(self):
        """function set_attr_sheet
        sets the attribute sheet
        returns attrs, the sheet with the attributes table
        """
        attrs = self.workbook.get_sheet_by_name("attributes")
        return attrs

    def save_file_changes(self):
        """function save_file_changes
        saves the emx file with the original file name
        """
        self.workbook.save(self.filename)
    
    def get_workbook(self):
        """function get_workbook
        returns the workbook to work with
        """
        return self.workbook
    
    def get_attr_sheet(self):
        """function get_attr_sheet
        returns the attributes sheet
        """
        return self.attr_sheet
    
    def get_genes_sheet(self):
        """function get_genes_sheet
        returns the genes sheet
        """
        return self.genes_sheet

    def alter_sheet(self, sheet, cell, value):
        """function alter_sheet
        This function can alter a sheet. It needs the sheet object to be altered,
        the cell which should be altered and the value that should be in the cell.
        """
        sheet[cell] = value

    def find_first_empty_cell(self, sheet):
        """function find_first_empty_cell
        This function finds the first empty cell in the given sheet (checks if the first column is empty)
        Returns the row number of the cell."""
        first_col = "A"
        iteration = 1
        empty_cell_not_found = True
        while empty_cell_not_found:
            if sheet[first_col+str(iteration)].value== None:
                empty_cell_not_found= False
            else:
                iteration += 1
        return iteration

